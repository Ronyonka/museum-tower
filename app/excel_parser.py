import datetime
from openpyxl import load_workbook
from .models import Tenant, TenantTransaction

def parse_excel_data(file_path):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        tenants = []
        tenant_transactions = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Check if the row represents a tenant
                tenant = Tenant(
                    name=row[0],
                    code=row[2],
                    main_unit_no=row[4],
                    property=row[6],
                    general_contact=row[8],
                    telephone=row[10],
                    lease_start_date=datetime.datetime.strptime(row[12], '%d/%m/%Y').date(),
                    lease_end_date=datetime.datetime.strptime(row[14], '%d/%m/%Y').date(),
                    vacate_date=datetime.datetime.strptime(row[16], '%d/%m/%Y').date() if row[16] else None,
                    active=True if not row[16] else False
                )
                tenants.append(tenant)

            if row[1] and row[3]:  # Check if the row represents a tenant transaction
                tenant_transaction = TenantTransaction(
                    tenant_id=tenant.id,  # Assuming the tenant object has an 'id' attribute
                    transaction_type=row[1],
                    transaction_date=datetime.datetime.strptime(row[3], '%d/%m/%Y').date(),
                    amount=float(row[7])
                )
                tenant_transactions.append(tenant_transaction)

        return tenants, tenant_transactions

    except Exception as e:
        print(f"Error parsing Excel data: {str(e)}")
        return None, None
